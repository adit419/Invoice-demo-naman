"""Generate realistic synthetic data for the commission-claim engine — at scale.

Outputs:
  data/volume_2026_05.csv   (Looker/Querybook-style aggregated volume export)
  data/commercials.xlsx     (rate cards / tiers / rebates the AR team maintains)
  data/emails.json          (commercials inbox the LLM scanner reads)

All data is fictional. Bank names are public; merchants are invented.
Edge cases are planted (2x of each type) so the engine visibly beats the manual
Excel process, and the emails are written as free text so a real LLM has to do
genuine extraction work to find the commercial signals.
"""
import json
import os
import random

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

random.seed(42)
OUT = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(OUT, exist_ok=True)

# --------------------------------------------------------------------------- #
# dimensions
# --------------------------------------------------------------------------- #
BANKS = [
    "HDFC", "ICICI", "Axis", "SBI", "Kotak", "YesBank",
    "PNB", "BoB", "IndusInd", "IDFC First", "Federal", "RBL",
]
MODES = ["RuPay Debit", "Credit Card", "Net Banking", "UPI", "Wallet"]
REGIONS = ["North", "South", "West", "East"]

PERIOD = "2026-05"
DAYS = list(range(1, 32))  # May has 31 days

# --- 120 invented merchants ---
_PREFIX = [
    "Zentro", "Bluepeak", "Nimbus", "Aether", "Saffron", "Vertex", "Lumen",
    "Coral", "Quanta", "Harbor", "Pixel", "Orchid", "Ironwood", "Solace",
    "Drift", "Maple", "Nova", "Verdant", "Tideline", "Cobalt", "Frostbyte",
    "Marigold", "Strata", "Echo", "Crimson", "Aurora", "Basalt", "Cedar",
    "Delta", "Ember", "Flint", "Granite", "Helio", "Indigo", "Juniper",
    "Kestrel", "Larch", "Monsoon", "Onyx", "Pyrite",
]
_SUFFIX = [
    "Retail", "Travels", "Grocers", "Mobility", "Foods", "Electronics",
    "Pharma", "Apparel", "Edutech", "Logistics", "Gaming", "Wellness",
    "Furniture", "Insurance", "Rentals", "Telecom", "Agro", "Resorts",
    "Auto", "Cloud", "Jewels", "Realty", "Media", "Fintech", "Labs",
    "Networks", "Studios", "Holdings", "Ventures", "Systems",
]
N_MERCHANTS = int(os.environ.get("N_MERCHANTS", "650"))  # knob for dataset size
MERCHANTS = []
_used = set()
_i = 1000
while len(MERCHANTS) < N_MERCHANTS:
    name = f"{random.choice(_PREFIX)} {random.choice(_SUFFIX)}"
    suffix = "" if name not in _used else f" {len(MERCHANTS)}"
    name = name + suffix
    _used.add(name)
    _i += 1
    MERCHANTS.append((f"M{_i}", name))

# size factor per merchant (drives GMV magnitude)
SIZE_CHOICES = [0.3, 0.5, 0.8, 1.0, 1.4, 2.0, 3.0]
sizes = {mid: random.choice(SIZE_CHOICES) for mid, _ in MERCHANTS}
regions = {mid: random.choice(REGIONS) for mid, _ in MERCHANTS}

MODE_SCALE = {"RuPay Debit": 1.0, "Credit Card": 1.4, "Net Banking": 2.2, "UPI": 0.6, "Wallet": 0.5}
AVG_TICKET = {"RuPay Debit": 1800, "Credit Card": 3200, "Net Banking": 6500, "UPI": 850, "Wallet": 600}


def merchant_routing(mid):
    """Each merchant routes through a subset of banks, each via a subset of modes."""
    nb = random.randint(2, 5)
    chosen_banks = random.sample(BANKS, nb)
    routing = []
    for b in chosen_banks:
        nm = random.randint(1, 3)
        for m in random.sample(MODES, nm):
            routing.append((b, m))
    return routing


rows = []
seq = 0
for mid, mname in MERCHANTS:
    region = regions[mid]
    for (bank, mode) in merchant_routing(mid):
        # daily batches this merchant pushes on this rail in the month (allow
        # several per day for larger merchants -> realistic high-volume export)
        n_batches = random.randint(6, 24)
        days = [random.choice(DAYS) for _ in range(n_batches)]
        for d in sorted(days):
            seq += 1
            base = sizes[mid] * random.uniform(0.7, 1.4)
            gmv = int(base * MODE_SCALE[mode] * random.uniform(8.0e5, 6.0e6))
            txn_count = max(1, int(gmv / (AVG_TICKET[mode] * random.uniform(0.8, 1.2))))
            batch_id = f"{bank[:3].upper()}-{PERIOD}-D{d:02d}-{seq:05d}"
            rows.append({
                "settlement_month": PERIOD, "batch_id": batch_id, "batch_day": d,
                "merchant_id": mid, "merchant_name": mname, "bank": bank,
                "payment_mode": mode, "region": region,
                "gmv": gmv, "txn_count": txn_count,
            })

df = pd.DataFrame(rows)


# --------------------------------------------------------------------------- #
# planted-volume helper
# --------------------------------------------------------------------------- #
def plant_volume(bank, mode, mid, mname, region, n, lo, hi, t_lo, t_hi):
    global seq
    for d in random.sample(DAYS, n):
        seq += 1
        df.loc[len(df)] = {
            "settlement_month": PERIOD, "batch_id": f"{bank[:3].upper()}-{PERIOD}-D{d:02d}-{seq:05d}",
            "batch_day": d, "merchant_id": mid, "merchant_name": mname, "bank": bank,
            "payment_mode": mode, "region": region,
            "gmv": int(random.uniform(lo, hi)), "txn_count": int(random.uniform(t_lo, t_hi)),
        }


# --- bias SBI Net Banking & PNB Net Banking upward so tiered slabs are crossed ---
for bk in ("SBI", "PNB"):
    m = (df.bank == bk) & (df.payment_mode == "Net Banking")
    df.loc[m, "gmv"] = (df.loc[m, "gmv"] * 3.0).astype(int)
    plant_volume(bk, "Net Banking", "M1004", "Aether Mobility", "South", 20, 8.0e7, 1.6e8, 9000, 18000)

# --- NO_TERM #1: YesBank UPI volume (no rate card row) ---
plant_volume("YesBank", "UPI", "M1011", "Pixel Gaming", "West", 14, 2.0e6, 5.0e6, 2500, 6000)
# --- NO_TERM #2: RBL Wallet volume (no rate card row) ---
plant_volume("RBL", "Wallet", "M1021", "Frostbyte Cloud", "North", 12, 1.5e6, 4.0e6, 2500, 6500)

# --- EXPIRED #1: Axis Credit Card (term expired 2026-04-30) ---
plant_volume("Axis", "Credit Card", "M1006", "Vertex Electronics", "South", 16, 4.0e6, 9.0e6, 1500, 3000)
# --- EXPIRED #2: IndusInd Net Banking (term expired 2026-03-31) ---
plant_volume("IndusInd", "Net Banking", "M1017", "Nova Telecom", "East", 14, 5.0e6, 1.0e7, 900, 1800)

# --- OVERLAP #1: HDFC RuPay Debit (v1 0.40 + v2 0.45 both active) ---
plant_volume("HDFC", "RuPay Debit", "M1003", "Nimbus Grocers", "North", 16, 5.0e6, 1.1e7, 3000, 6000)
# --- OVERLAP #2: ICICI Credit Card (v1 0.20 + v2 0.22 both active) ---
plant_volume("ICICI", "Credit Card", "M1008", "Coral Apparel", "West", 16, 6.0e6, 1.2e7, 2000, 4000)

# --- DUPLICATE #1 & #2: duplicate two existing batches exactly ---
for bank, mode in (("HDFC", "Credit Card"), ("SBI", "RuPay Debit")):
    sub = df[(df.bank == bank) & (df.payment_mode == mode)]
    if len(sub):
        df.loc[len(df)] = sub.iloc[0].to_dict()  # identical row incl. batch_id

df = df.sort_values(["bank", "payment_mode", "merchant_id", "batch_day"]).reset_index(drop=True)
df.to_csv(os.path.join(OUT, "volume_2026_05.csv"), index=False)
print(f"volume rows: {len(df):,}  | total GMV (Cr): {df.gmv.sum()/1e7:,.1f}")


# ============================ commercials.xlsx ============================
# Base rate per (bank, mode). PCT modes carry a %; per-txn modes carry rupees.
PCT_MODES = {"RuPay Debit", "Credit Card", "Wallet"}
BASE_PCT = {"RuPay Debit": 0.40, "Credit Card": 0.19, "Wallet": 0.45}
BASE_TXN = {"Net Banking": 9.0, "UPI": 0.28}
# small per-bank jitter so the card looks negotiated, not templated
BANK_JITTER = {b: random.uniform(-0.03, 0.04) for b in BANKS}
TXN_JITTER = {b: random.uniform(-1.0, 1.2) for b in BANKS}

# rails intentionally absent from the card (NO_TERM scenarios)
SKIP_TERMS = {("YesBank", "UPI"), ("RBL", "Wallet")}

rate_cards = []
_t = 0


def add_term(bank, mode, basis, rate_pct, fee_per_txn, eff_from, eff_to, version, status, notes):
    global _t
    _t += 1
    rate_cards.append([
        f"T{_t:03d}", bank, mode, "ALL", basis, rate_pct, fee_per_txn,
        eff_from, eff_to, version, status, notes,
    ])


for bank in BANKS:
    for mode in MODES:
        if (bank, mode) in SKIP_TERMS:
            continue
        # SBI / PNB Net Banking are tiered
        if mode == "Net Banking" and bank in ("SBI", "PNB"):
            add_term(bank, mode, "TIERED_GMV", None, None, "2026-01-01", None, 1, "ACTIVE", "See tiers sheet")
            continue
        if mode in PCT_MODES:
            rate = round(BASE_PCT[mode] + BANK_JITTER[bank], 2)
            add_term(bank, mode, "PCT_GMV", rate, None, "2026-01-01", None, 1, "ACTIVE", "")
        else:
            # UPI fees are a few paise; only Net Banking carries the larger rupee jitter
            jitter = BANK_JITTER[bank] if mode == "UPI" else TXN_JITTER[bank]
            fee = round(max(0.20, BASE_TXN[mode] + jitter), 2)
            add_term(bank, mode, "PER_TXN", None, fee, "2026-01-01", None, 1, "ACTIVE", "")

# ---- planted overlaps: a second, higher ACTIVE version effective this month ----
add_term("HDFC", "RuPay Debit", "PCT_GMV", 0.45, None, "2026-05-01", None, 2, "ACTIVE",
         "Renegotiated rate effective May — legacy v1 not retired in sheet")
add_term("ICICI", "Credit Card", "PCT_GMV", 0.22, None, "2026-05-01", None, 2, "ACTIVE",
         "Repriced CC effective May — older v1 still present")

# ---- planted expiries ----
add_term("Axis", "Credit Card", "PCT_GMV", 0.20, None, "2025-01-01", "2026-04-30", 1, "EXPIRED",
         "Lapsed 30 Apr — renewal pending")
add_term("IndusInd", "Net Banking", "PER_TXN", None, 8.5, "2024-06-01", "2026-03-31", 1, "EXPIRED",
         "Lapsed 31 Mar — renewal pending")

rc_cols = ["term_id", "bank", "payment_mode", "merchant_scope", "basis", "rate_pct", "fee_per_txn",
           "effective_from", "effective_to", "version", "status", "notes"]

# IndusInd Net Banking and Axis Credit Card should ONLY exist as their expired
# terms (renewal pending), so the bank renewal emails resolve a real EXPIRED
# exception in the worktable instead of silently billing a stale active rate.
rate_cards = [r for r in rate_cards
              if not (r[1] == "IndusInd" and r[2] == "Net Banking" and r[10] == "ACTIVE")
              and not (r[1] == "Axis" and r[2] == "Credit Card" and r[10] == "ACTIVE")]

# SBI RuPay Debit sits exactly at the 0.40% book rate so the bank's "rate remains
# 0.40%, unchanged" email reconciles as ALIGNED — the no-false-alarm case — rather
# than a spurious drift off a jittered 0.39%.
for r in rate_cards:
    if r[1] == "SBI" and r[2] == "RuPay Debit" and r[10] == "ACTIVE":
        r[5] = 0.40

tiers = [  # term_id, tier_no, min_gmv (incl), max_gmv (excl, blank=open), rate_pct
    # SBI Net Banking
    *[],
]
# attach tiers to whichever term ids ended up tiered
tier_terms = [r[0] for r in rate_cards if r[4] == "TIERED_GMV"]
for tid in tier_terms:
    tiers += [
        [tid, 1, 0, 2000000000, 0.30],
        [tid, 2, 2000000000, 6000000000, 0.25],
        [tid, 3, 6000000000, None, 0.20],
    ]
tier_cols = ["term_id", "tier_no", "min_gmv", "max_gmv", "rate_pct"]

rebates = [  # rebate_id, bank, mode, scope, rebate_pct, eff_from, eff_to, notes
    ["R001", "HDFC", "RuPay Debit", "ALL", 0.10, "2026-04-01", None, "Govt RuPay subvention"],
    ["R002", "ICICI", "RuPay Debit", "ALL", 0.10, "2026-04-01", None, "Govt RuPay subvention"],
    ["R003", "SBI", "RuPay Debit", "ALL", 0.10, "2026-04-01", None, "Govt RuPay subvention"],
    ["R004", "Axis", "RuPay Debit", "ALL", 0.10, "2026-04-01", None, "Govt RuPay subvention"],
    ["R005", "Kotak", "RuPay Debit", "ALL", 0.10, "2026-04-01", None, "Govt RuPay subvention"],
    ["R006", "PNB", "RuPay Debit", "ALL", 0.10, "2026-04-01", None, "Govt RuPay subvention"],
]
reb_cols = ["rebate_id", "bank", "payment_mode", "merchant_scope", "rebate_pct", "effective_from", "effective_to", "notes"]


# ---- styled workbook ----
wb = Workbook()
HEAD = Font(name="Calibri", bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="0B1F3A")
THIN = Side(style="thin", color="D9E1EA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_sheet(ws, cols, data, widths):
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEAD
        cell.fill = FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for r in data:
        ws.append(list(r))  # None -> genuinely empty cell, keeps numeric cols numeric
    for i in range(len(data) + 1):
        for c in range(1, len(cols) + 1):
            ws.cell(row=i + 1, column=c).border = BORDER
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"


ws1 = wb.active
ws1.title = "rate_cards"
write_sheet(ws1, rc_cols, rate_cards, [8, 11, 13, 14, 12, 9, 11, 13, 12, 8, 11, 46])
ws2 = wb.create_sheet("tiers")
write_sheet(ws2, tier_cols, tiers, [9, 9, 16, 16, 9])
ws3 = wb.create_sheet("rebates")
write_sheet(ws3, reb_cols, rebates, [9, 9, 13, 14, 11, 13, 12, 26])
ws4 = wb.create_sheet("readme")
ws4["A1"] = "Commercials workbook — maintained by AR/commercials team. Synthetic sample data."
ws4["A1"].font = Font(bold=True)
notes = [
    "", "basis values:",
    "  PCT_GMV     = fee = gmv * rate_pct%",
    "  PER_TXN     = fee = txn_count * fee_per_txn (INR)",
    "  TIERED_GMV  = progressive slabs on monthly bank+mode GMV (see tiers sheet)",
    "  REBATE_PCT  = additive rebate on gmv (rebates sheet)",
    "", "resolution: match bank+mode; specific merchant_scope beats ALL;",
    "effective_from <= period start and (effective_to blank or >= period start);",
    "if several valid, take highest version. Expired/overlapping terms are flagged.",
]
for i, n in enumerate(notes, 3):
    ws4[f"A{i}"] = n
ws4.column_dimensions["A"].width = 80

wb.save(os.path.join(OUT, "commercials.xlsx"))
print(f"rate-card terms: {len(rate_cards)}  | tiered terms: {len(tier_terms)}  | rebates: {len(rebates)}")
print("wrote commercials.xlsx and volume_2026_05.csv")


# ============================ emails.json ============================
# Free-text commercials inbox. The LLM scanner must extract bank / mode /
# proposed rate / effective date from messy prose, then code reconciles vs the
# rate card. Signal emails are crafted to land on specific statuses; the rest is
# realistic noise so extraction has to discriminate.
SENDERS = {
    # bank primary contacts
    "HDFC": ("Rohan Mehta", "rohan.mehta@hdfcbank.example"),
    "ICICI": ("Priya Nair", "priya.nair@icicibank.example"),
    "Axis": ("Karan Shah", "karan.shah@axisbank.example"),
    "SBI": ("Anil Kapoor", "anil.kapoor@sbi.example"),
    "Kotak": ("Sneha Iyer", "sneha.iyer@kotak.example"),
    "YesBank": ("Vikram Rao", "vikram.rao@yesbank.example"),
    "PNB": ("Deepa Menon", "deepa.menon@pnb.example"),
    "IndusInd": ("Arjun Reddy", "arjun.reddy@indusind.example"),
    "RBL": ("Meera Joshi", "meera.joshi@rblbank.example"),
    # bank secondary / treasury / legal contacts (loop in on real threads)
    "HDFC_treasury": ("Nikhil Verma", "nikhil.verma@hdfcbank.example"),
    "HDFC_ops": ("Sana Kapadia", "sana.kapadia@hdfcbank.example"),
    "Axis_legal": ("Ritu Desai", "ritu.desai@axisbank.example"),
    "IndusInd_ops": ("Kavya Nair", "kavya.nair@indusind.example"),
    "RBL_onboarding": ("Farhan Qureshi", "farhan.qureshi@rblbank.example"),
    "YesBank_ops": ("Ananya Bose", "ananya.bose@yesbank.example"),
    # internal AR / finance team members
    "internal": ("Finance Ops", "ar-team@neoflo.example"),
    "internal_lead": ("Aditya Rao", "aditya.rao@neoflo.example"),
    "internal_analyst": ("Fatima Sheikh", "fatima.sheikh@neoflo.example"),
}

# who is treated as "our side" — renders with the AR-team badge (direction outbound)
INTERNAL_KEYS = {"internal", "internal_lead", "internal_analyst"}

emails = []
_eid = 0
_tid = 0


def add_email(thread_id, subject, sender_key, direction, day, body, cc=None):
    global _eid
    _eid += 1
    name, addr = SENDERS[sender_key]
    emails.append({
        "id": f"EM-{_eid:03d}",
        "thread_id": thread_id,
        "subject": subject,
        "sender_name": name,
        "sender_email": addr,
        "direction": direction,
        "date": f"2026-05-{day:02d}",
        "cc": ", ".join(cc) if cc else "",
        "body": body,
    })


def new_thread():
    global _tid
    _tid += 1
    return f"TH-{_tid:02d}"


# Threads read like a real mailbox: a bank flags something, the AR team replies,
# colleagues / treasury / legal get looped in, and the exchange runs over several
# days before it settles. Extraction pulls a signal from each email; downstream
# code takes the LATEST-dated commercial email per (bank, mode) as the source of
# truth. So every thread below lands its final agreed rate on the last commercial
# message; the follow-ups after it are paperwork / acks with no new number.

# --- HERO 1: HDFC RuPay supersede thread (0.48, corrected to 0.50 -> DRIFT_UP vs card 0.45) ---
th = new_thread()
add_email(th, "RuPay Debit MDR revision — Neoflo programme", "HDFC", "inbound", 5,
          "Hi Aditya, team,\n\n"
          "Further to the QBR last week, we're revising the RuPay Debit MDR for the Neoflo "
          "programme to 0.48% with effect from 1 May 2026. I've looped in Nikhil from our "
          "Treasury desk who's driving the repricing. Kindly update your books and confirm "
          "once done.\n\n"
          "Best regards,\nRohan Mehta\nKey Accounts, HDFC Bank",
          cc=["Nikhil Verma", "Aditya Rao"])
add_email(th, "Re: RuPay Debit MDR revision — Neoflo programme", "internal", "outbound", 5,
          "Hi Rohan,\n\n"
          "Thanks for the note. Before we book it — just confirming this is 0.48% on RuPay "
          "Debit GMV, effective 1 May, superseding the current 0.45% on our card? Adding "
          "Fatima from my team who'll process the rate-card update.\n\n"
          "Regards,\nFinance Ops, Neoflo",
          cc=["Fatima Sheikh", "Nikhil Verma"])
add_email(th, "Re: RuPay Debit MDR revision — Neoflo programme", "HDFC", "inbound", 6,
          "Yes, 0.48% superseding 0.45%. Let me get Treasury to put the final number in "
          "writing before you lock it — Nikhil is validating against the network circular "
          "today.\n\n— Rohan",
          cc=["Nikhil Verma", "Fatima Sheikh"])
add_email(th, "Re: RuPay Debit MDR revision — Neoflo programme", "HDFC_treasury", "inbound", 8,
          "Nikhil here from HDFC Treasury. Still reconciling the RuPay slab against the "
          "revised interchange — I'll revert with the confirmed number by tomorrow. Please "
          "hold the update until then.\n\n"
          "Regards,\nNikhil Verma\nTreasury, HDFC Bank",
          cc=["Rohan Mehta"])
add_email(th, "Re: RuPay Debit MDR revision — Neoflo programme", "HDFC", "inbound", 9,
          "Apologies for the churn, team — correction on my earlier note. Treasury has now "
          "finalised RuPay Debit at 0.50%, not 0.48%. The 0.50% is the agreed number, "
          "effective 1st May 2026. Please book against this.\n\n— Rohan",
          cc=["Nikhil Verma", "Aditya Rao", "Fatima Sheikh"])
add_email(th, "Re: RuPay Debit MDR revision — Neoflo programme", "internal", "outbound", 9,
          "Understood Rohan — booking RuPay Debit at 0.50% effective 1 May 2026 and "
          "retiring the older rate-card versions. Fatima will push the update today.\n\n"
          "— Finance Ops",
          cc=["Fatima Sheikh", "Nikhil Verma"])
add_email(th, "Re: RuPay Debit MDR revision — Neoflo programme", "HDFC_ops", "inbound", 12,
          "Hi all, Sana from HDFC ops here. Sharing the signed pricing addendum for your "
          "records (attached). This is just the paperwork for what Rohan confirmed — no "
          "change to the agreed terms. Do reach me for any settlement queries going "
          "forward.\n\n"
          "Regards,\nSana Kapadia\nHDFC Bank",
          cc=["Rohan Mehta", "Aditya Rao"])

# --- HERO 2: Axis CC renewal (0.42 -> resolves EXPIRED) ---
th = new_thread()
add_email(th, "Axis Credit Card — renewal confirmation", "Axis", "inbound", 3,
          "Hello,\n\n"
          "Confirming the renewal of the Axis Credit Card programme for Neoflo. The renewed "
          "rate is 0.42% on credit card GMV, effective 1 May 2026. The prior agreement "
          "lapsed at the end of April; this renewal supersedes it. Copying Ritu from Legal "
          "who will send the renewal paperwork.\n\n"
          "Regards,\nKaran Shah\nPartnerships, Axis Bank",
          cc=["Ritu Desai", "Aditya Rao"])
add_email(th, "Re: Axis Credit Card — renewal confirmation", "internal", "outbound", 3,
          "Thanks Karan — to confirm, the previous Axis CC term lapsed 30 Apr and 0.42% "
          "applies from 1 May. We'd flagged this as an expired term on our side, so this "
          "clears it. Ritu, please send the addendum when ready.\n\n— Finance Ops",
          cc=["Ritu Desai", "Fatima Sheikh"])
add_email(th, "Re: Axis Credit Card — renewal confirmation", "Axis_legal", "inbound", 6,
          "Hi, Ritu from Axis Legal. Attaching the countersigned renewal addendum for the "
          "Credit Card programme — commercial terms are exactly as Karan set out, nothing "
          "new to action. Please retain for your records.\n\n"
          "Best,\nRitu Desai\nLegal, Axis Bank",
          cc=["Karan Shah"])
add_email(th, "Re: Axis Credit Card — renewal confirmation", "internal", "outbound", 7,
          "Received, thanks Ritu. Filed against the renewal. Marking the expired term "
          "cleared on our end.\n\n— Finance Ops")

# --- HERO 3: YesBank UPI new term (Rs 0.30/txn -> resolves NO_TERM) ---
th = new_thread()
add_email(th, "Go-live: YesBank UPI pricing for Neoflo", "YesBank", "inbound", 4,
          "Team,\n\n"
          "We're live on UPI for Neoflo as of this week. Pricing is Rs 0.30 per "
          "transaction, w.e.f 1 May 2026. There's no prior arrangement — this is the "
          "first UPI term between us. Ananya from our ops team (cc'd) will share the "
          "daily settlement MIS. Let me know once it's reflected your side.\n\n"
          "Thanks,\nVikram Rao\nYes Bank",
          cc=["Ananya Bose", "Aditya Rao"])
add_email(th, "Re: Go-live: YesBank UPI pricing for Neoflo", "internal", "outbound", 4,
          "Great news Vikram. Adding YesBank UPI at Rs 0.30/txn effective 1 May. Since "
          "there was no term on file we'd been holding this volume un-billable — good to "
          "get it in. Will confirm once the rate card is updated.\n\n— Finance Ops",
          cc=["Fatima Sheikh", "Ananya Bose"])
add_email(th, "Re: Go-live: YesBank UPI pricing for Neoflo", "YesBank_ops", "inbound", 7,
          "Hi, Ananya from YesBank ops. First week's UPI settlement MIS is on the shared "
          "SFTP. Purely informational — no change to the Rs 0.30/txn we agreed. Flag any "
          "reconciliation gaps and I'll chase them.\n\n"
          "Regards,\nAnanya Bose\nYes Bank",
          cc=["Vikram Rao"])

# --- HERO 4: RBL Wallet new pricing (0.55% -> resolves NO_TERM) ---
th = new_thread()
add_email(th, "RBL Wallet — commercial terms", "RBL", "inbound", 8,
          "Hi,\n\n"
          "Pleased to share the agreed commercials for the RBL Wallet integration: 0.55% "
          "of wallet GMV, effective from 1 May 2026. This is a fresh arrangement — there "
          "was no prior term on file. Looping in Farhan who'll run the onboarding.\n\n"
          "Best,\nMeera Joshi\nRBL Bank",
          cc=["Farhan Qureshi", "Aditya Rao"])
add_email(th, "Re: RBL Wallet — commercial terms", "internal", "outbound", 9,
          "Thanks Meera — confirming RBL Wallet at 0.55% from 1 May. This volume was "
          "sitting un-billable with no rate on file, so this is exactly what we needed. "
          "Fatima is updating the rate card now.\n\n— Finance Ops",
          cc=["Fatima Sheikh", "Farhan Qureshi"])
add_email(th, "Re: RBL Wallet — commercial terms", "RBL_onboarding", "inbound", 11,
          "Hello, Farhan from RBL onboarding. Sharing the integration checklist and test "
          "credentials — this is technical only, the 0.55% commercial term stands as Meera "
          "confirmed. Let's target production cutover next week.\n\n"
          "Regards,\nFarhan Qureshi\nRBL Bank",
          cc=["Meera Joshi"])

# --- HERO 5: IndusInd Net Banking renewal (Rs 9.00/txn -> resolves EXPIRED) ---
th = new_thread()
add_email(th, "IndusInd Net Banking — renewal", "IndusInd", "inbound", 5,
          "Hello team,\n\n"
          "The Net Banking arrangement that expired on 31 March has now been renewed. The "
          "new fee is Rs 9.00 per transaction, effective 1 May 2026. Please bring this "
          "back into billing. Kavya from ops (cc'd) will send the signed letter.\n\n"
          "Regards,\nArjun Reddy\nIndusInd Bank",
          cc=["Kavya Nair", "Aditya Rao"])
add_email(th, "Re: IndusInd Net Banking — renewal", "internal", "outbound", 5,
          "Thanks Arjun — we'd parked IndusInd Net Banking as an expired term and held the "
          "claim. Bringing it back into billing at Rs 9.00/txn from 1 May.\n\n— Finance Ops",
          cc=["Fatima Sheikh", "Kavya Nair"])
add_email(th, "Re: IndusInd Net Banking — renewal", "IndusInd_ops", "inbound", 8,
          "Hi, Kavya from IndusInd ops. Attaching the signed renewal letter for your file "
          "— it mirrors the Rs 9.00/txn Arjun confirmed, no new terms. Reach me for "
          "settlement queries.\n\n"
          "Regards,\nKavya Nair\nIndusInd Bank",
          cc=["Arjun Reddy"])

# --- ICICI CC reduction (0.18 vs card 0.22 -> DRIFT_DOWN, over-billing risk) ---
th = new_thread()
add_email(th, "ICICI Credit Card — rate reduction", "ICICI", "inbound", 8,
          "Hi,\n\n"
          "Following the volume review, ICICI is reducing the Credit Card MDR for Neoflo "
          "to 0.18% effective 1 May 2026. Please ensure you are not billing the older, "
          "higher rate from this month onward.\n\n"
          "— Priya Nair\nICICI Bank",
          cc=["Aditya Rao"])
add_email(th, "Re: ICICI Credit Card — rate reduction", "internal", "outbound", 8,
          "Noted Priya. We'll stop applying the prior rate from the May cycle and switch "
          "to the reduced rate. Flagging internally so we don't over-bill on the changeover "
          "— Fatima, please prioritise this one.\n\n— Finance Ops",
          cc=["Fatima Sheikh"])
add_email(th, "Re: ICICI Credit Card — rate reduction", "internal_analyst", "outbound", 12,
          "Confirming I've staged the ICICI CC change so the May cycle bills at the reduced "
          "rate, not the old card rate. No further input needed from ICICI — this is an "
          "internal note.\n\n— Fatima Sheikh, Neoflo",
          cc=["Aditya Rao"])

# --- SBI RuPay aligned (0.40 == card -> ALIGNED, no false alarm) ---
th = new_thread()
add_email(th, "SBI RuPay Debit — confirmation of existing rate", "SBI", "inbound", 3,
          "Hi,\n\n"
          "Just confirming for your records that the SBI RuPay Debit rate remains 0.40%, "
          "unchanged for the May cycle. No action needed from your end.\n\n"
          "Regards,\nAnil Kapoor\nState Bank of India",
          cc=["Aditya Rao"])
add_email(th, "Re: SBI RuPay Debit — confirmation of existing rate", "internal", "outbound", 4,
          "Thanks Anil — confirmed, no change on our side either. Our card already reads "
          "0.40%, so nothing to update. Closing this out.\n\n— Finance Ops")

# --- Kotak UPI drift up (Rs 0.32 vs card ~0.28 -> DRIFT_UP) ---
th = new_thread()
add_email(th, "Kotak UPI — revised per-txn fee", "Kotak", "inbound", 10,
          "Team,\n\n"
          "Please note the Kotak UPI per-transaction fee is being revised to Rs 0.32 with "
          "effect from 1 May 2026, in line with the recent network changes. Circular "
          "attached.\n\n"
          "Regards,\nSneha Iyer\nKotak Mahindra Bank",
          cc=["Aditya Rao"])
add_email(th, "Re: Kotak UPI — revised per-txn fee", "internal", "outbound", 10,
          "Thanks Sneha — noted, we'll move Kotak UPI to Rs 0.32/txn from 1 May. Our card "
          "currently sits lower, so we'll add the revised term.\n\n— Finance Ops",
          cc=["Fatima Sheikh"])
add_email(th, "Re: Kotak UPI — revised per-txn fee", "Kotak", "inbound", 13,
          "Thanks for the quick turnaround. The network circular I referenced is attached "
          "again for completeness — no change to the Rs 0.32 already confirmed.\n\n— Sneha")

# --- PNB RuPay drift down (0.36 vs card ~0.40 -> DRIFT_DOWN) ---
th = new_thread()
add_email(th, "PNB RuPay Debit — concessional rate", "PNB", "inbound", 12,
          "Hi,\n\n"
          "As discussed on our call, PNB is offering a concessional RuPay Debit rate of "
          "0.36% for Neoflo, effective this month. Please update your records.\n\n"
          "— Deepa Menon\nPunjab National Bank",
          cc=["Aditya Rao"])
add_email(th, "Re: PNB RuPay Debit — concessional rate", "internal", "outbound", 12,
          "Thanks Deepa — to be precise for our books, can you confirm the effective date "
          "is 1 May 2026? We'll lower the RuPay Debit rate to 0.36% accordingly.\n\n"
          "— Finance Ops")
add_email(th, "Re: PNB RuPay Debit — concessional rate", "PNB", "inbound", 14,
          "Confirmed — 0.36% on RuPay Debit, effective 1 May 2026. Please go ahead and "
          "update.\n\n— Deepa Menon, Punjab National Bank")

# --- HDFC Credit Card drift up (0.21 vs card 0.19 -> small DRIFT_UP) ---
th = new_thread()
add_email(th, "HDFC Credit Card — minor revision", "HDFC", "inbound", 13,
          "Minor update: the HDFC Credit Card MDR moves to 0.21% from 1 May 2026. It's "
          "marginal, but flagging so the books stay in sync.\n\n— Rohan Mehta, HDFC Bank",
          cc=["Sana Kapadia"])
add_email(th, "Re: HDFC Credit Card — minor revision", "internal", "outbound", 13,
          "Thanks Rohan — noted, we'll nudge HDFC Credit Card to 0.21% from 1 May.\n\n"
          "— Finance Ops",
          cc=["Fatima Sheikh"])
add_email(th, "Re: HDFC Credit Card — minor revision", "HDFC_ops", "inbound", 15,
          "Sana here — acknowledging on Rohan's behalf, the 0.21% is correct. Nothing "
          "further needed.\n\n— Sana Kapadia, HDFC Bank")

# --- Federal Wallet: internal thread that still carries one clean signal (0.50%) ---
th = new_thread()
add_email(th, "Catch-up + Federal Wallet pricing", "internal_lead", "outbound", 9,
          "Folks,\n\n"
          "Couple of things from the Federal call. They were happy with the ramp and want "
          "to push volume. On commercials, Federal confirmed Wallet at 0.50% effective "
          "1 May (it was verbal earlier, now in writing). They'll also send the UPI MIS by "
          "Friday and want a roadmap review next month.\n\n"
          "Let's sync Monday.\n\n— Aditya Rao",
          cc=["Fatima Sheikh", "Finance Ops"])
add_email(th, "Re: Catch-up + Federal Wallet pricing", "internal", "outbound", 10,
          "Thanks Aditya. I'll add Federal Wallet at 0.50% from 1 May to the rate card and "
          "reconcile the un-billed volume. Will pick up the UPI MIS separately once it "
          "lands.\n\n— Finance Ops",
          cc=["Fatima Sheikh"])

# --------- NOISE EMAILS (no commercial change) ---------
NOISE = [
    ("internal", "Settlement file for April — shared", 2,
     "Hi all, the April settlement file is on the shared drive. Numbers tally with the "
     "bank statements. Note HDFC processed roughly 0.4 million transactions last month — "
     "good growth. No pricing changes here, just FYI. — AR Team"),
    ("SBI", "Festive offer — cashback campaign", 14,
     "Hello, SBI is running a festive cashback of up to Rs 500 for cardholders in May. "
     "This is a consumer promo and does not affect your MDR. Sharing for awareness. — Anil"),
    ("Kotak", "Scheduled downtime notice", 15,
     "Team, Kotak net banking will have a maintenance window on 18 May, 1-3 AM. No "
     "commercial impact. Regards, Sneha"),
    ("ICICI", "KYC documentation request", 16,
     "Hi, we need updated KYC documents for the Neoflo entity for our annual refresh. "
     "Please share the certificate of incorporation. Nothing to do with pricing. — Priya"),
    ("internal", "Re: QBR deck", 17,
     "Attached the QBR deck draft. Slide 7 has the GMV split. We grew 12% MoM. Let's "
     "review before sending. — AR Team"),
    ("Axis", "Holiday calendar 2026", 18,
     "Sharing the Axis settlement holiday calendar for the rest of 2026 so you can plan "
     "cut-offs. No rate changes. — Karan"),
    ("HDFC", "Contact update", 19,
     "Please note my colleague Sana will be your secondary contact going forward. Reach "
     "either of us for ops queries. — Rohan"),
    ("internal", "Reminder: month-end reconciliation", 20,
     "Reminder to close the May reconciliation by the 5th. Flagging that YesBank UPI and "
     "RBL Wallet still need rate-card entries — chasing the banks separately. — AR Team"),
    ("PNB", "Webinar invite — digital payments", 21,
     "You're invited to our digital payments webinar on 25 May. Topics include UPI trends. "
     "Purely informational. — Deepa"),
    ("YesBank", "Thank you note", 22,
     "Thanks for a smooth go-live this week. Looking forward to scaling UPI volumes "
     "together. — Vikram"),
    ("ICICI", "Quarterly business review — scheduling", 23,
     "Can we lock the QBR for the first week of June? Proposing 3rd or 4th. Agenda to "
     "follow. — Priya"),
    ("internal", "FYI: new merchant onboarded", 24,
     "Onboarded Pyrite Systems this week, routing via HDFC and Axis. Standard rates apply, "
     "nothing bespoke. — AR Team"),
    ("Kotak", "Survey: partner satisfaction", 25,
     "We'd value your feedback in our annual partner survey. Takes 5 minutes. No pricing "
     "content. — Sneha"),
    ("Federal", "Intro — Federal relationship manager", 26,
     "Hello, introducing myself as your new RM at Federal. Happy to help with any ops or "
     "commercial questions. — Federal Team"),
    ("internal", "Re: settlement variance Apr", 27,
     "The Rs 2 lakh variance in April was a timing difference, now squared off. No action. "
     "— AR Team"),
]
for sk, subj, day, body in NOISE:
    if sk == "Federal":
        SENDERS.setdefault("Federal", ("Federal Team", "rm@federalbank.example"))
    add_email(new_thread(), subj, sk, "inbound" if sk != "internal" else "internal", day, body)

mailbox = {
    "mailbox": "ar-team@neoflo.example",
    "period": PERIOD,
    "emails": emails,
}
with open(os.path.join(OUT, "emails.json"), "w") as f:
    json.dump(mailbox, f, indent=2)
print(f"emails: {len(emails)}  ({_tid} threads)")


# ============================ SQLite store ============================
# Persist inputs + computed results so the API serves huge data a page at a
# time via SQL. The engine remains the calculation source of truth.
import time as _time  # noqa: E402

import compute_claim as _engine  # noqa: E402
import dbstore as _store  # noqa: E402

_rc_df = pd.DataFrame(rate_cards, columns=rc_cols)
_tiers_df = pd.DataFrame(tiers, columns=tier_cols)
_reb_df = pd.DataFrame(rebates, columns=reb_cols)
for _d in (_rc_df, _reb_df):
    for _c in ("effective_from", "effective_to"):
        _d[_c] = pd.to_datetime(_d[_c], errors="coerce")

print("writing SQLite inputs…")
_store.write_inputs(df, _rc_df, _tiers_df, _reb_df, emails)

print("computing claim (one-time build)…")
_t0 = _time.time()
_res = _engine.compute(df, _rc_df, _tiers_df, _reb_df)
print(f"  compute: {_time.time() - _t0:.1f}s  | firm claim Rs {_res['grand_total']/1e7:,.2f} Cr "
      f"| billable {_res['billable_lines']:,} | exceptions {len(_res['exceptions'])}")
_store.write_results(_res, PERIOD)
print(f"wrote {_store.DB}")
